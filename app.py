import streamlit as st
import pandas as pd
import io
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLUNAS_PADRAO = ['DATA_VENDA', 'DATA_PAGAMENTO', 'TIPO_VENDA', 'VALOR BRUTO', 'VALOR TAXA']


def formatar_moeda(valor):
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def extrair_info_empresa(arquivo_getnet):
    empresa = None
    periodo = None
    try:
        xls = pd.ExcelFile(arquivo_getnet)
        for sheet in xls.sheet_names:
            raw = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=6)
            for r in range(raw.shape[0]):
                for c in range(raw.shape[1]):
                    val = raw.iat[r, c]
                    if not isinstance(val, str):
                        continue
                    if empresa is None and val.strip().startswith('Razão Social:'):
                        empresa = val.split(':', 1)[1].strip()
                    if periodo is None and val.strip().startswith('Periodo:'):
                        periodo = val.split(':', 1)[1].strip()
            if empresa and periodo:
                break
    except Exception:
        pass
    return empresa, periodo


def competencia_de_datas(serie_datas):
    s = pd.to_datetime(serie_datas, errors='coerce').dropna()
    if s.empty:
        return None
    moda = s.dt.to_period('M').mode()
    if moda.empty:
        return None
    return moda.iloc[0].strftime('%m/%Y')


def periodo_texto_para_competencia(periodo_texto):
    if not periodo_texto:
        return None
    primeira_data = periodo_texto.split(' a ')[0].strip()
    dt = pd.to_datetime(primeira_data, dayfirst=True, errors='coerce')
    if pd.isna(dt):
        return None
    return dt.strftime('%m/%Y')


def processar_ifood(arquivo_ifood):
    df_ifood = pd.read_excel(arquivo_ifood)
    colunas = set(df_ifood.columns)

    # Identifica o nome exato da coluna de lançamento no Modelo 1
    col_lancamento = None
    for nome in ['descricao_lancamento', 'tipo_lancamento', 'Tipo de lançamento']:
        if nome in colunas:
            col_lancamento = nome
            break

    # MODELO 1: Extrato Financeiro
    colunas_m1 = {'fato_gerador', 'valor', 'data_criacao_pedido_associado', 'data_repasse_esperada'}
    if colunas_m1.issubset(colunas) and col_lancamento:
        vendas = df_ifood[df_ifood['fato_gerador'].astype(str).str.strip().str.upper() == 'VENDA'].copy()
        vendas = vendas[vendas[col_lancamento].astype(str).str.strip().str.upper() == 'ENTRADA FINANCEIRA']
        
        vendas['DATA_VENDA'] = pd.to_datetime(vendas['data_criacao_pedido_associado'], errors='coerce').dt.date
        vendas['DATA_PAGAMENTO'] = pd.to_datetime(vendas['data_repasse_esperada'], errors='coerce').dt.date
        vendas['VALOR BRUTO'] = pd.to_numeric(vendas['valor'], errors='coerce').fillna(0)
        
        bruto_ifood = vendas.groupby(['DATA_VENDA', 'DATA_PAGAMENTO'], dropna=False)['VALOR BRUTO'].sum().reset_index()
        bruto_ifood['VALOR TAXA'] = 0.00
        bruto_ifood['TIPO_VENDA'] = 'IFOOD'
        
        return bruto_ifood[COLUNAS_PADRAO], None

    # MODELO 2: Relatório de Pedidos
    colunas_m2 = {'DATA E HORA DO PEDIDO', 'STATUS FINAL DO PEDIDO', 'VALOR DOS ITENS (R$)'}
    if colunas_m2.issubset(colunas):
        vendas = df_ifood[df_ifood['STATUS FINAL DO PEDIDO'].astype(str).str.strip().str.upper().str.contains('CONCLU')].copy()
        
        vendas['DATA_VENDA'] = pd.to_datetime(vendas['DATA E HORA DO PEDIDO'], errors='coerce').dt.date
        vendas['DATA_PAGAMENTO'] = pd.NaT  
        vendas['VALOR BRUTO'] = pd.to_numeric(vendas['VALOR DOS ITENS (R$)'], errors='coerce').fillna(0)
        
        bruto_ifood = vendas.groupby(['DATA_VENDA', 'DATA_PAGAMENTO'], dropna=False)['VALOR BRUTO'].sum().reset_index()
        bruto_ifood['VALOR TAXA'] = 0.00
        bruto_ifood['TIPO_VENDA'] = 'IFOOD'
        
        nome_loja = None
        if 'NOME DA LOJA' in colunas:
            modas = df_ifood['NOME DA LOJA'].dropna()
            if not modas.empty:
                nome_loja = modas.mode().iloc[0]
                
        return bruto_ifood[COLUNAS_PADRAO], nome_loja

    raise ValueError("A planilha anexada não possui a estrutura exata do Modelo 1 ou Modelo 2 do iFood informados.")


st.set_page_config(page_title="Automação Contábil", layout="centered")
st.title("📊 Consolidador Getnet e iFood")
st.write("Anexe a planilha da GETNET, do IFOOD, ou as duas (do mesmo período) para gerar a consolidação contábil diária.")

arquivo_getnet = st.file_uploader("📂 Anexe a planilha da GETNET (opcional)", type=['xlsx'])
arquivo_ifood = st.file_uploader("📂 Anexe a planilha do IFOOD (opcional)", type=['xlsx'])

if 'dados_processados' not in st.session_state:
    st.session_state['dados_processados'] = False
if 'buffer_excel' not in st.session_state:
    st.session_state['buffer_excel'] = None

if st.button("🚀 Processar Dados"):
    if arquivo_getnet is None and arquivo_ifood is None:
        st.warning("⚠️ Anexe ao menos uma planilha (Getnet e/ou iFood) para processar.")
    else:
        try:
            with st.spinner('Engrenagens girando... Processando matriz de dados! ⚙️'):
                df_vazio = pd.DataFrame(columns=COLUNAS_PADRAO)

                empresa_nome_getnet = None
                periodo_getnet_texto = None
                competencia_getnet = None

                if arquivo_getnet is not None:
                    empresa_nome_getnet, periodo_getnet_texto = extrair_info_empresa(arquivo_getnet)

                    # Cartões
                    df_cartoes = pd.read_excel(arquivo_getnet, sheet_name='CARTÕES', header=7)
                    df_cartoes = df_cartoes[df_cartoes['STATUS DA TRANSAÇÃO'] == 'Aprovada'].copy()
                    df_cartoes['DATA_VENDA'] = pd.to_datetime(df_cartoes['DATA/HORA DA VENDA'], dayfirst=True, errors='coerce').dt.date
                    df_cartoes['DATA_PAGAMENTO'] = pd.to_datetime(df_cartoes['DATA PREVISTA DO 1º PAGAMENTO'], dayfirst=True, errors='coerce').dt.date
                    df_cartoes['VALOR BRUTO'] = pd.to_numeric(df_cartoes['VALOR BRUTO'], errors='coerce').fillna(0)
                    df_cartoes['VALOR TAXA'] = pd.to_numeric(df_cartoes['VALOR TAXA'], errors='coerce').fillna(0)
                    df_cartoes['TIPO_VENDA'] = 'Getnet ' + df_cartoes['BANDEIRA'].astype(str) + ' ' + df_cartoes['MODALIDADE'].astype(str)
                    df_cartoes_limpo = df_cartoes[COLUNAS_PADRAO]

                    # PIX
                    df_pix = pd.read_excel(arquivo_getnet, sheet_name='PIX', header=7)
                    df_pix = df_pix[df_pix['STATUS'] == 'Paga'].copy()
                    df_pix['DATA_VENDA'] = pd.to_datetime(df_pix['DATA/HORA DA VENDA'], dayfirst=True, errors='coerce').dt.date
                    df_pix['DATA_PAGAMENTO'] = df_pix['DATA_VENDA']
                    df_pix['VALOR BRUTO'] = pd.to_numeric(df_pix['VALOR DA VENDA'], errors='coerce').fillna(0)
                    df_pix['VALOR TAXA'] = pd.to_numeric(df_pix['VALOR TAXA'], errors='coerce').fillna(0)
                    df_pix['TIPO_VENDA'] = 'Getnet PIX'
                    df_pix_limpo = df_pix[COLUNAS_PADRAO]

                    # Vouchers
                    df_voucher = pd.read_excel(arquivo_getnet, sheet_name='VOUCHER', header=7)
                    df_voucher = df_voucher[df_voucher['STATUS'] == 'Aprovada'].copy()
                    df_voucher['DATA_VENDA'] = pd.to_datetime(df_voucher['DATA DA VENDA'], dayfirst=True, errors='coerce').dt.date
                    df_voucher['DATA_PAGAMENTO'] = df_voucher['DATA_VENDA']
                    df_voucher['VALOR BRUTO'] = pd.to_numeric(df_voucher['VALOR DA VENDA'], errors='coerce').fillna(0)
                    df_voucher['VALOR TAXA'] = 0.00
                    df_voucher['TIPO_VENDA'] = df_voucher['BANDEIRA'].replace({'Sodexo': 'Pluxee'})
                    df_voucher_limpo = df_voucher[COLUNAS_PADRAO]

                    df_getnet_consolidado = pd.concat([df_cartoes_limpo, df_pix_limpo, df_voucher_limpo], ignore_index=True)
                    getnet_agrupado = df_getnet_consolidado.groupby(['DATA_VENDA', 'DATA_PAGAMENTO', 'TIPO_VENDA'], dropna=False)[['VALOR BRUTO', 'VALOR TAXA']].sum().reset_index()

                    competencia_getnet = periodo_texto_para_competencia(periodo_getnet_texto) or competencia_de_datas(getnet_agrupado['DATA_VENDA'])
                else:
                    df_cartoes_limpo = df_pix_limpo = df_voucher_limpo = df_vazio.copy()
                    getnet_agrupado = df_vazio.copy()

                nome_loja_ifood = None
                competencia_ifood = None

                if arquivo_ifood is not None:
                    ifood_agrupado, nome_loja_ifood = processar_ifood(arquivo_ifood)
                    competencia_ifood = competencia_de_datas(ifood_agrupado['DATA_VENDA'])
                else:
                    ifood_agrupado = df_vazio.copy()

                if arquivo_getnet is not None and arquivo_ifood is not None:
                    if competencia_getnet and competencia_ifood and competencia_getnet != competencia_ifood:
                        st.error(
                            f"🚫 As planilhas são de competências diferentes e não podem ser processadas juntas.\n\n"
                            f"- Getnet: **{competencia_getnet}**\n"
                            f"- iFood: **{competencia_ifood}**\n\n"
                            f"Anexe relatórios do mesmo mês/ano, ou processe um de cada vez."
                        )
                        st.stop()

                empresa_nome = empresa_nome_getnet or nome_loja_ifood
                competencia = competencia_getnet or competencia_ifood

                # Consolidação
                diario_geral = pd.concat([getnet_agrupado, ifood_agrupado], ignore_index=True)
                diario_geral = diario_geral.sort_values(by=['DATA_VENDA', 'DATA_PAGAMENTO', 'TIPO_VENDA']).reset_index(drop=True)

                # Regras de Negócio e Formatação para OFX Externo
                diario_geral['TIPO_VENDA'] = diario_geral['TIPO_VENDA'].astype(str).str.upper()
                diario_geral['VALOR TAXA'] = diario_geral['VALOR TAXA'].abs()
                
                # Transformação de datas
                diario_geral['DATA_VENDA'] = diario_geral['DATA_VENDA'].apply(
                    lambda x: x.strftime('%d/%m/%Y') if pd.notnull(x) else ''
                )
                diario_geral['DATA_PAGAMENTO'] = diario_geral['DATA_PAGAMENTO'].apply(
                    lambda x: x.strftime('%d.%m.%Y') if pd.notnull(x) else ''
                )

                # Geração do Resumo
                resumo_geral = diario_geral.groupby('TIPO_VENDA')[['VALOR BRUTO', 'VALOR TAXA']].sum().reset_index()
                linha_total = pd.DataFrame([{'TIPO_VENDA': 'TOTAL GERAL', 'VALOR BRUTO': resumo_geral['VALOR BRUTO'].sum(), 'VALOR TAXA': resumo_geral['VALOR TAXA'].sum()}])
                resumo_geral = pd.concat([resumo_geral, linha_total], ignore_index=True)

                # Escrita no Buffer
                buffer = io.BytesIO()
                linha_inicio_resumo = 3 if (empresa_nome or competencia) else 0

                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    diario_geral.to_excel(writer, sheet_name='Movimento_Diario', index=False)
                    resumo_geral.to_excel(writer, sheet_name='Resumo_Totais', index=False, startrow=linha_inicio_resumo)

                    workbook = writer.book
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill("solid", fgColor="2F4F4F")
                    info_font = Font(bold=True, color="2F4F4F")
                    border = Border(left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                                    top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3'))

                    if linha_inicio_resumo:
                        ws_resumo = workbook['Resumo_Totais']
                        ws_resumo.cell(row=1, column=1, value=f"Empresa: {empresa_nome or 'Não identificada'}").font = info_font
                        ws_resumo.cell(row=2, column=1, value=f"Competência: {competencia or 'Não identificada'}").font = info_font

                    linha_cabecalho = {'Movimento_Diario': 1, 'Resumo_Totais': linha_inicio_resumo + 1}

                    for sheet_name in workbook.sheetnames:
                        worksheet = workbook[sheet_name]
                        cab = linha_cabecalho.get(sheet_name, 1)
                        for cell in worksheet[cab]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        for row in worksheet.iter_rows(min_row=cab + 1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                            for cell in row:
                                cell.border = border
                                if isinstance(cell.value, float) or isinstance(cell.value, int):
                                    cell.number_format = '#,##0.00'
                        for col in worksheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except: pass
                            worksheet.column_dimensions[column].width = (max_length + 2)

                # Persistência no Session State
                st.session_state['buffer_excel'] = buffer.getvalue()
                st.session_state['resumo_dados'] = {
                    'empresa': empresa_nome or 'Não identificada',
                    'competencia': competencia or 'Não identificada',
                    'total_getnet': df_cartoes_limpo['VALOR BRUTO'].sum() + df_pix_limpo['VALOR BRUTO'].sum() if arquivo_getnet else 0,
                    'total_ifood': ifood_agrupado['VALOR BRUTO'].sum() if arquivo_ifood else 0,
                    'total_vouchers': df_voucher_limpo['VALOR BRUTO'].sum() if arquivo_getnet else 0,
                    'arquivo_getnet': arquivo_getnet is not None,
                    'arquivo_ifood': arquivo_ifood is not None,
                    'ifood_sem_data': arquivo_ifood is not None and ifood_agrupado['DATA_PAGAMENTO'].isna().all(),
                    'vouchers_agrupados': df_voucher_limpo.groupby('TIPO_VENDA')['VALOR BRUTO'].sum() if arquivo_getnet else None
                }
                st.session_state['dados_processados'] = True

        except Exception as e:
            st.error(f"🚨 Erro na estrutura dos arquivos: {e}")

if st.session_state.get('dados_processados'):
    resumo = st.session_state['resumo_dados']
    total_geral = resumo['total_getnet'] + resumo['total_ifood'] + resumo['total_vouchers']
    
    st.success("✨ Processamento concluído com sucesso!")
    st.subheader("📋 Resumo Operacional Bruto")
    st.markdown(f"🏢 **Empresa:** {resumo['empresa']}  \n📅 **Competência:** {resumo['competencia']}")
    
    col1, col2 = st.columns(2)
    with col1:
        if resumo['arquivo_getnet']:
            st.markdown(f"💳 **Getnet (Cartões + PIX):** {formatar_moeda(resumo['total_getnet'])}")
        if resumo['arquivo_ifood']:
            st.markdown(f"🍔 **iFood:** {formatar_moeda(resumo['total_ifood'])}")
        st.markdown(f"💰 **TOTAL GERAL:** {formatar_moeda(total_geral)}")
    
    with col2:
        if resumo['arquivo_getnet'] and resumo['vouchers_agrupados'] is not None and not resumo['vouchers_agrupados'].empty:
            st.markdown("🎟️ **Vouchers (Detalhado):**")
            for bandeira, valor in resumo['vouchers_agrupados'].items():
                st.markdown(f"- **{bandeira}:** {formatar_moeda(valor)}")
                
    if resumo['ifood_sem_data']:
        st.caption("ℹ️ O modelo de planilha do iFood anexado não informa data de repasse/liquidação — a coluna DATA_PAGAMENTO ficou em branco para essas linhas.")

    st.markdown("---")

    st.download_button(
        label="⬇️ Baixar Planilha Consolidada",
        data=st.session_state['buffer_excel'],
        file_name="Consolidado_Contabilidade.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
